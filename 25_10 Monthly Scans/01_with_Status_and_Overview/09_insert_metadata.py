import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from copy import copy

# Add parent directory (where config.py is) to sys.path to import config
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, '..'))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

import config  # Import config.py from parent directory

# Configuration
base_dir = os.getcwd()
final_reports_dir = base_dir  # Final reports in current directory
metadata_dir = os.path.join(base_dir, 'All Metadata')  # Metadata folder

countries = ['Brazil', 'CSA', 'India', 'Mexico', 'SFTL']
report_types = ['Network', 'Server']

def copy_cell_style(source_cell, target_cell):
    """Copy style properties from source_cell to target_cell."""
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = copy(source_cell.number_format)
    target_cell.protection = copy(source_cell.protection)

def insert_metadata_into_final_report(final_report_path, metadata_file_path):
    print(f"\nProcessing final report: {final_report_path.name}")
    print(f"Using metadata file: {metadata_file_path.name}")
    
    # Load workbooks
    final_wb = load_workbook(final_report_path)
    metadata_wb = load_workbook(metadata_file_path)
    
    # Use active sheets (adjust if needed)
    final_ws = final_wb.active
    metadata_ws = metadata_wb.active
    
    metadata_rows = metadata_ws.max_row
    metadata_cols = metadata_ws.max_column
    
    print(f"Metadata size: {metadata_rows} rows x {metadata_cols} columns")
    
    # Insert empty rows at top of final report to shift data down
    final_ws.insert_rows(1, amount=metadata_rows)
    
    # Copy metadata cell values and styles
    for row in range(1, metadata_rows + 1):
        for col in range(1, metadata_cols + 1):
            source_cell = metadata_ws.cell(row=row, column=col)
            target_cell = final_ws.cell(row=row, column=col, value=source_cell.value)
            copy_cell_style(source_cell, target_cell)
    
    # Copy row heights from metadata to final report
    for row in range(1, metadata_rows + 1):
        height = metadata_ws.row_dimensions[row].height
        if height is not None:
            final_ws.row_dimensions[row].height = height
    
    # Insert one extra empty row after metadata to push header to row 8
    final_ws.insert_rows(metadata_rows + 1, amount=1)
    
    # Save the updated final report
    final_wb.save(final_report_path)
    print(f"Metadata inserted successfully into {final_report_path.name}")

def main():
    for country in countries:
        for report_type in report_types:
            # Use year and current_month_num from config.py instead of hardcoding
            final_report_filename = f'{country}_{config.year}_{config.current_month_num}_Final_Report_with_Status_{report_type}_Report.xlsx'
            metadata_filename = f'{country}_{report_type}_Metadata.xlsx'
            
            final_report_path = Path(final_reports_dir) / final_report_filename
            metadata_file_path = Path(metadata_dir) / metadata_filename
            
            if final_report_path.exists() and metadata_file_path.exists():
                insert_metadata_into_final_report(final_report_path, metadata_file_path)
            else:
                print(f"Missing file(s): {final_report_path.name} or {metadata_file_path.name}")

if __name__ == "__main__":
    main()
