import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

# Add parent directory (where config.py is located) to sys.path
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, '..'))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

import config  # Import config.py from parent directory

# Use config variables instead of hardcoded ones
previous_month = config.previous_month
current_month = config.current_month
previous_month_num = config.previous_month_num
current_month_num = config.current_month_num
year = config.year

countries = ['Brazil', 'CSA', 'India', 'Mexico', 'SFTL']
report_types = ['Network', 'Server']

def create_summary_table(country, top_n_ips=5):
    summary_data = []
    
    for report_type in report_types:
        file_path = f'{country}_{year}_{current_month_num}_Final_Report_with_Status_{report_type}_Report.xlsx'
        if os.path.exists(file_path):
            workbook = pd.ExcelFile(file_path)
            sheet_name = [sheet for sheet in workbook.sheet_names if sheet != 'Overview'][0]
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=7)
            
            severity_levels = {
                5: 'Critical',
                4: 'High',
                3: 'Medium',
                2: 'Low'
            }
            
            top_issues = []
            missing_severities = []
            
            for severity, severity_name in severity_levels.items():
                issues = df[df['Severity'] == severity]
                if not issues.empty:
                    top_issues = issues['Title'].value_counts().head(4).index.tolist()
                    if len(top_issues) < 4:
                        for lower_severity in range(severity - 1, 1, -1):
                            lower_issues = df[df['Severity'] == lower_severity]
                            if not lower_issues.empty:
                                top_issues += lower_issues['Title'].value_counts().head(4 - len(top_issues)).index.tolist()
                                if len(top_issues) >= 4:
                                    break
                    break
                else:
                    missing_severities.append(severity_name)
            
            if missing_severities:
                message = f"We don't have {', '.join(missing_severities)} issues in this month's scan. Issues of lower severity are mentioned below:"
            else:
                message = ""
            
            ip_details = df[df['Title'].isin(top_issues)]['IP'].value_counts().head(top_n_ips).index.tolist()
            ip_details_str = ', '.join(ip_details)
            
            summary_data.append({
                'Asset Type': f'{report_type} IP',
                'Title': '• ' + '\n• '.join(top_issues) if not message else message + '\n• ' + '\n• '.join(top_issues),
                'IP with more vulnerabilities': ip_details_str
            })
        else:
            print(f'File {file_path} does not exist.')
    
    return summary_data

def save_summary_to_excel(summary_data, output_file):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    
    worksheet.merge_cells('D5:F5')
    worksheet['D5'] = "Summary of top critical issues:"
    worksheet['D5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    worksheet['D6'] = "Asset Type (Network/Server IP)"
    worksheet['E6'] = "Vulnerability Name"
    worksheet['F6'] = "Asset/IP detail with more vulnerabilities"
    
    worksheet.column_dimensions['D'].width = 25
    worksheet.column_dimensions['E'].width = 60
    worksheet.column_dimensions['F'].width = 20
    
    for col in ['D', 'E', 'F']:
        worksheet[f'{col}6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    row = 7
    for data in summary_data:
        worksheet[f'D{row}'] = data['Asset Type']
        worksheet[f'E{row}'] = data['Title']
        worksheet[f'F{row}'] = data['IP with more vulnerabilities']
        
        worksheet[f'E{row}'].alignment = Alignment(wrap_text=True)
        worksheet[f'F{row}'].alignment = Alignment(wrap_text=True)
        
        row += 1
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row_cells in worksheet.iter_rows(min_row=5, max_row=row-1, min_col=4, max_col=6):
        for cell in row_cells:
            cell.border = thin_border
    
    workbook.save(output_file)

for country in countries:
    summary_data = create_summary_table(country, top_n_ips=5)
    output_file = f'{country}_{year}_{current_month_num}_Summary_Report.xlsx'
    save_summary_to_excel(summary_data, output_file)
    print(f'Summary report saved to {output_file}')
