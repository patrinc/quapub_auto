import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import math

# Add parent directory (where config.py is located) to sys.path for import
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, '..'))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

import config

# Use config variables instead of hardcoded ones
current_date = datetime.now().date()
fixed_start_date = config.fixed_start_date  # assumed datetime.date object in config.py
countries = ['Brazil', 'CSA', 'India', 'Mexico', 'SFTL']
report_types = ['Network', 'Server']
year = config.year
current_month_num = config.current_month_num
header_row = 8

sla_days = {
    5: 7,
    4: 15,
    3: 21,
    2: 30,
    1: 30
}

non_breach_statuses = {"Patched", "Mitigated", "False Positive", "Not Applicable", "Risk Accepted", "Inactive", "Shutdown"}
breach_statuses = {"Unpatched", "In Progress", "Deferred", ""}

base_folder = Path(fr"C:\Users\example.user1\OneDrive - exampledomain\Assignments\VAPT\Infra Scanning\25_{current_month_num} Monthly Scans\01_with_Status_and_Overview\SLA Status")

country_folders = {
    "Brazil": "Brazil Qualys Monthly Scans",
    "CSA": "CSA Qualys Monthly Scans",
    "India": "India Qualys Monthly Scans",
    "Mexico": "Mexico Qualys Monthly Scans",
    "SFTL": "SFTL Qualys Monthly Scans"
}

severity_levels = [
    'Severity - 5 (Critical)',
    'Severity - 4 (High)',
    'Severity - 3 (Medium)',
    'Severity - 2 (Low)',
    'Severity - 1 (Info)'
]

severity_colors = {
    'Severity - 5 (Critical)': 'FFC00000',
    'Severity - 4 (High)': 'FFFF0000',
    'Severity - 3 (Medium)': 'FFFFC000',
    'Severity - 2 (Low)': 'FF92D050',
    'Severity - 1 (Info)': 'FF00B0F0',
}

status_fill_colors = {
    "Patched": "FFA5D6A7",
    "Unpatched": "FFEF9A9A",
    "Risk Accepted": "FFFFF59D",
    "Mitigated": "FF64B5F6",
    "In Progress": "FF90CAF9",
    "False Positive": "FFE0E0E0",
    "Not Applicable": "FF9575CD",
    "Deferred": "FFFFF176",
    "Inactive": "FFB0BEC5",
    "Shutdown": "FF78909C"
}

def set_border_around_range(ws, min_row, max_row, min_col, max_col, border):
    for col in range(min_col, max_col + 1):
        top_cell = ws.cell(row=min_row, column=col)
        top_cell.border = Border(top=border.top, left=top_cell.border.left, right=top_cell.border.right, bottom=top_cell.border.bottom)
        bottom_cell = ws.cell(row=max_row, column=col)
        bottom_cell.border = Border(bottom=border.bottom, left=bottom_cell.border.left, right=bottom_cell.border.right, top=bottom_cell.border.top)
    for row in range(min_row, max_row + 1):
        left_cell = ws.cell(row=row, column=min_col)
        left_cell.border = Border(left=border.left, top=left_cell.border.top, right=left_cell.border.right, bottom=left_cell.border.bottom)
        right_cell = ws.cell(row=row, column=max_col)
        right_cell.border = Border(right=border.right, top=right_cell.border.top, left=right_cell.border.left, bottom=right_cell.border.bottom)

def unmerge_cells_in_range(ws, min_row, max_row, min_col, max_col):
    to_unmerge = []
    for merged_range in ws.merged_cells.ranges:
        mr_min_row, mr_max_row = merged_range.min_row, merged_range.max_row
        mr_min_col, mr_max_col = merged_range.min_col, merged_range.max_col
        if not (mr_max_row < min_row or mr_min_row > max_row or mr_max_col < min_col or mr_min_col > max_col):
            to_unmerge.append(merged_range.coord)
    for coord in to_unmerge:
        try:
            ws.unmerge_cells(coord)
        except KeyError:
            pass

def add_status_severity_placeholder_table(overview_sheet):
    heading_row = 10
    start_col = 5  # E
    end_col = start_col + len(severity_levels)

    overview_sheet.merge_cells(start_row=heading_row, start_column=start_col, end_row=heading_row, end_column=end_col)
    heading_cell = overview_sheet.cell(row=heading_row, column=start_col)
    heading_cell.value = "Overview: Counts of Status per Severity"
    heading_cell.font = Font(bold=True)
    heading_cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
    set_border_around_range(overview_sheet, heading_row, heading_row, start_col, end_col, thin_border)

    start_row = 11
    status_col = 5
    severity_start_col = 6
    severity_end_col = severity_start_col + len(severity_levels) - 1
    last_status_row = start_row + len(status_fill_colors)

    status_header_cell = overview_sheet.cell(row=start_row, column=status_col, value="Status Value")
    status_header_cell.font = Font(bold=True, color="FFFFFFFF")
    status_header_cell.alignment = Alignment(horizontal='center', vertical='center')
    status_header_cell.fill = PatternFill(start_color="FF284464", end_color="FF284464", fill_type='solid')

    for i, severity in enumerate(severity_levels):
        cell = overview_sheet.cell(row=start_row, column=severity_start_col + i, value=severity)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        fill_color = severity_colors.get(severity)
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    for i, status in enumerate(status_fill_colors.keys(), start=start_row + 1):
        cell = overview_sheet.cell(row=i, column=status_col, value=status)
        cell.font = Font(bold=False)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        fill_color = status_fill_colors[status]
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    overview_sheet.column_dimensions['D'].width = 20
    for col_letter in ['E', 'F', 'G', 'H', 'I', 'J']:
        overview_sheet.column_dimensions[col_letter].width = 18

    thin_border = Border(left=Side(border_style="thin", color="000000"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="000000"),
                         bottom=Side(border_style="thin", color="000000"))

    for row in range(start_row, last_status_row + 1):
        for col in range(status_col, severity_end_col + 1):
            cell = overview_sheet.cell(row=row, column=col)
            cell.border = thin_border

def get_due_severities(elapsed_days):
    due_sevs = []
    for sev, days in sorted(sla_days.items(), reverse=True):
        if elapsed_days > days:
            due_sevs.append(sev)
    return due_sevs

def check_sla_breach(severity, status, start_date, due_severities):
    if severity not in due_severities:
        return None
    if severity not in sla_days:
        return True
    if status is None or (isinstance(status, float) and math.isnan(status)):
        status_norm = ""
    else:
        status_norm = str(status).strip().title()
    if status_norm in {s.title() for s in non_breach_statuses}:
        return False
    if status_norm == "":
        return True
    if status_norm in {s.title() for s in breach_statuses}:
        if start_date is None:
            return True
        allowed_days = sla_days[severity]
        start_ts = pd.Timestamp(start_date)
        current_ts = pd.Timestamp(current_date)
        due_date = start_ts + pd.Timedelta(days=allowed_days)
        if current_ts > due_date:
            return f"SLA Breached (Due: {due_date.strftime('%Y-%m-%d')})"
        else:
            return False
    return False

def update_sla_breach_column(workbook, df, header_row=header_row):
    sheet_name = df.attrs.get('sheet_name')
    worksheet = workbook[sheet_name]

    headers = [cell.value for cell in worksheet[header_row]]

    try:
        it_col_idx = headers.index("IT") + 1
    except ValueError:
        it_col_idx = None

    try:
        status_col_idx = headers.index("Status") + 1
    except ValueError:
        status_col_idx = None

    if "SLA Status" in headers:
        sla_col_idx = headers.index("SLA Status") + 1
    else:
        if it_col_idx is not None:
            sla_col_idx = it_col_idx + 1  # After IT column
        elif status_col_idx is not None:
            sla_col_idx = status_col_idx + 1  # After Status column
        else:
            sla_col_idx = len(headers) + 1  # Append at the end

        # Set header for SLA Status column
        worksheet.cell(row=header_row, column=sla_col_idx, value="SLA Status")

    header_cell = worksheet.cell(row=header_row, column=sla_col_idx)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_cell.alignment = Alignment(vertical='top', horizontal='center')
    header_cell.font = Font(bold=True)
    header_cell.border = thin_border

    red_fill = PatternFill(start_color="FFFF6F61", end_color="FFFF6F61", fill_type="solid")

    for idx, row in df.iterrows():
        excel_row = header_row + 1 + idx
        sla_breached = row.get('SLA Status')

        cell = worksheet.cell(row=excel_row, column=sla_col_idx)

        if isinstance(sla_breached, str):  # message string
            cell.value = sla_breached
            cell.fill = red_fill
            cell.font = Font(bold=True)
        elif sla_breached is True:
            cell.value = "Overdue"
            cell.fill = red_fill
            cell.font = Font(bold=True)
        else:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)  # Clear fill
            cell.font = Font(bold=False)

def find_column_letters(ws, header_row=8):
    col_map = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value:
            col_map[str(cell.value).strip().lower()] = get_column_letter(col_idx)
    return col_map

def write_overview_status_severity_formulas(overview_sheet, scan_results_sheet_name, scan_results_ws, header_row=11):
    status_col = 5
    severity_start_col = 6
    statuses = list(status_fill_colors.keys())
    severity_map = {
        'Severity - 5 (Critical)': 5,
        'Severity - 4 (High)': 4,
        'Severity - 3 (Medium)': 3,
        'Severity - 2 (Low)': 2,
        'Severity - 1 (Info)': 1,
    }

    col_map = find_column_letters(scan_results_ws, header_row=8)
    severity_col_letter = col_map.get('severity')
    status_col_letter = col_map.get('status')

    if not severity_col_letter or not status_col_letter:
        print(f"Could not find 'Severity' or 'Status' columns in Scan Results sheet!")
        return

    for row_offset, status in enumerate(statuses, start=header_row + 1):
        for idx, sev_label in enumerate(severity_levels):
            col = severity_start_col + idx
            sev_num = severity_map[sev_label]
            formula = f'''=COUNTIFS('{scan_results_sheet_name}'!${severity_col_letter}:${severity_col_letter}, {sev_num}, '{scan_results_sheet_name}'!${status_col_letter}:${status_col_letter}, "{status}")'''
            cell = overview_sheet.cell(row=row_offset, column=col)
            cell.value = formula
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=False)

def write_sla_breach_summary_static(overview_sheet, df, header_row=11, gap_rows=2):
    breached_df = df[df['SLA Status'].apply(lambda x: isinstance(x, str) or x is True)]

    start_col = 5  # E

    if breached_df.empty:
        start_row = header_row + len(status_fill_colors) + gap_rows + 1
        overview_sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 5)
        cell = overview_sheet.cell(row=start_row, column=start_col)
        cell.value = "No SLA breach has been identified."
        cell.font = Font(bold=True, color="FF000000")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        return

    breached_df = breached_df.copy()
    breached_df.loc[:, 'Status_for_summary'] = breached_df['Status'].replace({"": "(Blank Status)"})

    all_sev_nums = [int(s.split()[2]) for s in severity_levels]
    sev_num_to_label = {int(s.split()[2]): s for s in severity_levels}

    pivot = breached_df.pivot_table(index='Status_for_summary', columns='Severity', values='SLA Status', aggfunc='count', fill_value=0)

    for sev in all_sev_nums:
        if sev not in pivot.columns:
            pivot[sev] = 0
    pivot = pivot[sorted(all_sev_nums, reverse=True)]

    pivot = pivot[pivot.sum(axis=1) > 0]

    start_row = header_row + len(status_fill_colors) + gap_rows + 1
    end_col = start_col + len(severity_levels)

    num_statuses = len(pivot.index)
    num_severities = len(severity_levels)
    min_row = start_row
    max_row = start_row + 1 + num_statuses
    min_col = start_col
    max_col = end_col

    # Unmerge all merged cells overlapping the SLA Breach Summary table range
    unmerge_cells_in_range(overview_sheet, min_row, max_row, min_col, max_col)

    overview_sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    heading_cell = overview_sheet.cell(row=start_row, column=start_col)
    heading_cell.value = "SLA Breach Summary"
    heading_cell.font = Font(bold=True)
    heading_cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
    set_border_around_range(overview_sheet, start_row, start_row, start_col, end_col, thin_border)

    header_row_num = start_row + 1
    overview_sheet.cell(row=header_row_num, column=start_col, value="Status Value").font = Font(bold=True, color="FFFFFFFF")
    overview_sheet.cell(row=header_row_num, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
    overview_sheet.cell(row=header_row_num, column=start_col).fill = PatternFill(start_color="FF284464", end_color="FF284464", fill_type='solid')
    overview_sheet.cell(row=header_row_num, column=start_col).border = thin_border

    for i, sev_label in enumerate(severity_levels):
        col_idx = start_col + 1 + i
        cell = overview_sheet.cell(row=header_row_num, column=col_idx, value=sev_label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        fill_color = severity_colors.get(sev_label)
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        cell.border = thin_border

    data_start_row = header_row_num + 1
    for row_offset, status in enumerate(pivot.index):
        status_cell = overview_sheet.cell(row=data_start_row + row_offset, column=start_col, value=status)
        status_cell.font = Font(bold=False)
        status_cell.alignment = Alignment(horizontal='left', vertical='center')
        fill_color = status_fill_colors.get(status)
        if status == "(Blank Status)":
            status_cell.fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type='solid')
        elif fill_color:
            status_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        status_cell.border = thin_border

        for i, sev_label in enumerate(severity_levels):
            sev_num = int(sev_label.split()[2])
            col_idx = start_col + 1 + i
            count = pivot.at[status, sev_num] if sev_num in pivot.columns else 0
            count_cell = overview_sheet.cell(row=data_start_row + row_offset, column=col_idx, value=count)
            count_cell.alignment = Alignment(horizontal='center', vertical='center')
            count_cell.font = Font(bold=False)
            count_cell.border = thin_border

    overview_sheet.column_dimensions[get_column_letter(start_col)].width = 20
    for i in range(len(severity_levels)):
        overview_sheet.column_dimensions[get_column_letter(start_col + 1 + i)].width = 18

def process_file(file_path, sheet_name, start_date):
    print(f"Processing: {file_path.name}")
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row - 1, engine='openpyxl')
        df.columns = df.columns.str.strip()
    except Exception as e:
        print(f"Failed to read {file_path.name}: {e}")
        return
    required_cols = ['Severity', 'Status']
    for col in required_cols:
        if col not in df.columns:
            print(f"Missing required column '{col}' in {file_path.name}")
            return

    if df.empty:
        print(f"No data found in sheet {sheet_name} of file {file_path.name}, skipping SLA calculation.")
        return

    df['Severity'] = pd.to_numeric(df['Severity'], errors='coerce').fillna(0).astype(int)
    df['Status'] = df['Status'].fillna("").astype(str).str.strip().str.title()
    elapsed_days = (current_date - start_date).days
    due_severities = get_due_severities(elapsed_days)
    print(f"Elapsed calendar days since start date: {elapsed_days}")
    print(f"Severity levels currently overdue for SLA breach check: {due_severities}")
    def compute_sla(row):
        sla_result = check_sla_breach(row['Severity'], row['Status'], start_date, due_severities)
        if row['Status'].title() == "Patched" and sla_result is False:
            return None
        return sla_result
    df['SLA Status'] = df.apply(compute_sla, axis=1)
    breached_count = df['SLA Status'].apply(lambda x: isinstance(x, str) or x is True).sum()
    df.attrs['sheet_name'] = sheet_name
    workbook = load_workbook(file_path)
    scan_results_ws = workbook[sheet_name]
    update_sla_breach_column(workbook, df, header_row=header_row)
    if 'Overview' in workbook.sheetnames:
        overview_sheet = workbook['Overview']
    else:
        overview_sheet = workbook.create_sheet('Overview')
        workbook._sheets.insert(0, workbook._sheets.pop(workbook._sheets.index(overview_sheet)))
    if overview_sheet.cell(row=11, column=5).value != "Status Value":
        add_status_severity_placeholder_table(overview_sheet)
    write_overview_status_severity_formulas(overview_sheet, sheet_name, scan_results_ws, header_row=11)
    write_sla_breach_summary_static(overview_sheet, df, header_row=11, gap_rows=2)
    workbook.save(file_path)
    print(f"SLA breaches identified: {breached_count}\n")

if __name__ == "__main__":
    for country in countries:
        folder_name = country_folders.get(country)
        if not folder_name:
            print(f"No folder configured for country: {country}")
            continue
        folder_path = base_folder / folder_name
        if not folder_path.exists():
            print(f"Folder not found for {country}: {folder_path}")
            continue
        for report_type in report_types:
            pattern = f"{country}_{year}_{current_month_num}_Final_Report_{report_type}_Report.xlsx"
            matched_files = list(folder_path.glob(pattern))
            if not matched_files:
                print(f"No matching report found for {country} {report_type} in {folder_path}")
                continue
            for file_path in matched_files:
                sheet_name = f"{country}_{report_type}_Scan_Result"
                process_file(file_path, sheet_name, fixed_start_date)
