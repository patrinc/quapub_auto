import os
import sys
import re
import shutil
from openpyxl import load_workbook

# Add parent directory (where config.py is located) to sys.path
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, '..'))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

import config  # Import config.py from parent directory

# Variables from config
year = config.year
current_month_num = config.current_month_num
countries = ['Brazil', 'CSA', 'India', 'Mexico', 'SFTL']
report_types = ['Network', 'Server']

# Compile the filename matching pattern:
pattern = re.compile(
    rf"({'|'.join(countries)})_{year}_{current_month_num}_Final_Report_with_Status_({'|'.join(report_types)})_Report\.xlsx"
)

# Columns to delete exactly by header name found in the first row
COLUMNS_TO_DELETE = [
    'Comments',
    'Status_prev',
    'Ticket No._prev',
    'Comments_prev',
    'IT_prev',
    'SLA Status_prev',
    'composite_key_no_port'
]

# Target directory with dynamic month folder
target_directory = rf"C:\Users\example.user1\OneDrive - exampledomain\Assignments\VAPT\Infra Scanning\25_{current_month_num} Monthly Scans\02_Final Reports_without_Overview"

if not os.path.exists(target_directory):
    os.makedirs(target_directory)

def remove_with_status_from_filename(filename):
    return filename.replace('_with_Status', '')

def delete_columns_from_excel(file_path, columns_to_delete):
    wb = load_workbook(file_path)
    for ws in wb.worksheets:
        # Read headers from the first row
        headers = [cell.value for cell in ws[1]]
        # Find columns to delete by their header names (1-based indexing)
        cols_to_delete = [idx + 1 for idx, header in enumerate(headers) if header in columns_to_delete]
        # Sort descending to avoid shifting issues when deleting multiple columns
        cols_to_delete.sort(reverse=True)
        for col_idx in cols_to_delete:
            ws.delete_cols(col_idx)
    wb.save(file_path)
    print(f"Deleted specified columns in '{os.path.basename(file_path)}'")

def main():
    # Iterate over files in current directory
    for file_name in os.listdir('.'):
        if pattern.match(file_name):
            print(f"Found matching file: {file_name}")
            # Copy file to target directory with renamed filename
            src_path = os.path.abspath(file_name)
            copied_file_name = remove_with_status_from_filename(file_name)
            dst_path = os.path.join(target_directory, copied_file_name)
            shutil.copy2(src_path, dst_path)
            print(f"Copied and renamed file to: {dst_path}")
            # Delete the specified columns in the copied file
            delete_columns_from_excel(dst_path, COLUMNS_TO_DELETE)

if __name__ == '__main__':
    main()
