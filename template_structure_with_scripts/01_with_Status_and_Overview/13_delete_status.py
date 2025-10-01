import os
import sys
from pathlib import Path
from copy import copy  
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

# Include config from parent directory
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, '..'))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

import config 

previous_month_num = config.previous_month_num
current_month_num = config.current_month_num
year = config.year

countries = ['Brazil', 'CSA', 'India', 'Mexico', 'SFTL']
report_types = ['Network', 'Server']

approval_folder = Path('For Approval')
approval_folder.mkdir(exist_ok=True)

HEADER_ROW = 8  

status_descriptions = {
    "Patched": "Vulnerability has been fixed with a patch.",
    "Unpatched": "Vulnerability is still present and unaddressed.",
    "Risk Accepted": "Risk is acknowledged but accepted without remediation.",
    "Mitigated": "Risk has been reduced by compensating controls.",
    "In Progress": "Remediation work is currently underway.",
    "False Positive": "Vulnerability identified but deemed invalid.",
    "Not Applicable": "Vulnerability does not apply to this system.",
    "Deferred": "Remediation postponed to a later date.",
    "Inactive": "Server is inactive and not currently in use.",
    "Shutdown": "Server has been shut down and is not operational."
}

allowed_status_values = list(status_descriptions.keys())

status_fill_colors = {
    "Patched": "A5D6A7",
    "Unpatched": "EF9A9A",
    "Risk Accepted": "FFF59D",
    "Mitigated": "64B5F6",
    "In Progress": "90CAF9",
    "False Positive": "E0E0E0",
    "Not Applicable": "9575CD",
    "Deferred": "FFF176",
    "Inactive": "B0BEC5",
    "Shutdown": "78909C"
}


def find_status_column_index(worksheet, header_row=HEADER_ROW):
    for cell in worksheet[header_row]:
        if cell.value and cell.value.strip().lower() == "status":
            return cell.column
    return None


def insert_ticket_no_column(worksheet, status_col_idx, header_row=HEADER_ROW):
    worksheet.insert_cols(status_col_idx + 1)
    new_col_idx = status_col_idx + 1
    header_cell = worksheet.cell(row=header_row, column=new_col_idx, value="Ticket No.")

    ref_cell = worksheet.cell(row=header_row, column=status_col_idx)

    header_cell.font = copy(ref_cell.font)
    header_cell.fill = copy(ref_cell.fill)
    header_cell.alignment = copy(ref_cell.alignment)
    header_cell.border = copy(ref_cell.border)

    for row in range(header_row + 1, worksheet.max_row + 1):
        worksheet.cell(row=row, column=new_col_idx).value = None


def insert_comments_column(worksheet, ticket_col_idx, header_row=HEADER_ROW):
    worksheet.insert_cols(ticket_col_idx + 1)
    new_col_idx = ticket_col_idx + 1
    header_cell = worksheet.cell(row=header_row, column=new_col_idx, value="Comments")

    ref_cell = worksheet.cell(row=header_row, column=ticket_col_idx)

    header_cell.font = copy(ref_cell.font)
    header_cell.fill = copy(ref_cell.fill)
    header_cell.alignment = copy(ref_cell.alignment)
    header_cell.border = copy(ref_cell.border)

    for row in range(header_row + 1, worksheet.max_row + 1):
        worksheet.cell(row=row, column=new_col_idx).value = None


def insert_it_column(worksheet, comments_col_idx, header_row=HEADER_ROW):
    worksheet.insert_cols(comments_col_idx + 1)
    new_col_idx = comments_col_idx + 1
    header_cell = worksheet.cell(row=header_row, column=new_col_idx, value="IT")

    ref_cell = worksheet.cell(row=header_row, column=comments_col_idx)

    header_cell.font = copy(ref_cell.font)
    header_cell.fill = copy(ref_cell.fill)
    header_cell.alignment = copy(ref_cell.alignment)
    header_cell.border = copy(ref_cell.border)

    for row in range(header_row + 1, worksheet.max_row + 1):
        worksheet.cell(row=row, column=new_col_idx).value = None


def clear_status_column_values(worksheet, status_col_idx, header_row=HEADER_ROW):
    for row in range(header_row + 1, worksheet.max_row + 1):
        worksheet.cell(row=row, column=status_col_idx).value = None


def add_status_data_validation_safe(worksheet, status_col_letter, allowed_values, header_row=HEADER_ROW):
    start_row = header_row + 1
    end_row = worksheet.max_row
    if end_row >= start_row:
        cell_range = f"{status_col_letter}{start_row}:{status_col_letter}{end_row}"
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(allowed_values)}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Entry",
            error="Your entry is not in the allowed list. Please select a value from the dropdown."
        )
        worksheet.add_data_validation(dv)
        dv.add(cell_range)
    else:
        print(f"Skipping status data validation for sheet '{worksheet.title}': no data rows.")


def add_ticket_no_numeric_validation_safe(worksheet, ticket_col_letter, header_row=HEADER_ROW):
    start_row = header_row + 1
    end_row = worksheet.max_row
    if end_row >= start_row:
        cell_range = f"{ticket_col_letter}{start_row}:{ticket_col_letter}{end_row}"
        dv = DataValidation(
            type="whole",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Entry",
            error="Please enter a numeric ticket number."
        )
        worksheet.add_data_validation(dv)
        dv.add(cell_range)
    else:
        print(f"Skipping Ticket No. data validation for sheet '{worksheet.title}': no data rows.")


def add_it_data_validation_safe(worksheet, it_col_letter, bu_name, header_row=HEADER_ROW):
    start_row = header_row + 1
    end_row = worksheet.max_row
    if end_row >= start_row:
        cell_range = f"{it_col_letter}{start_row}:{it_col_letter}{end_row}"
        allowed_values = [bu_name, "GSS"]
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(allowed_values)}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Entry",
            error="Please select a value from the dropdown."
        )
        worksheet.add_data_validation(dv)
        dv.add(cell_range)
    else:
        print(f"Skipping IT data validation for sheet '{worksheet.title}': no data rows.")


def add_status_conditional_formatting_safe(worksheet, status_col_letter, header_row=HEADER_ROW):
    start_row = header_row + 1
    end_row = worksheet.max_row
    if end_row >= start_row:
        cell_range = f"{status_col_letter}{start_row}:{status_col_letter}{end_row}"
        for status, color in status_fill_colors.items():
            formula = f'=${status_col_letter}{start_row}="{status}"'
            fill = PatternFill(start_color=f"FF{color}", end_color=f"FF{color}", fill_type="solid")
            rule = FormulaRule(formula=[formula], fill=fill)
            worksheet.conditional_formatting.add(cell_range, rule)
    else:
        print(f"Skipping status conditional formatting for sheet '{worksheet.title}': no data rows.")


def add_status_legend_sheet(workbook, status_descriptions):
    if "Status Description" in workbook.sheetnames:
        workbook.remove(workbook["Status Description"])
    legend_sheet = workbook.create_sheet(title="Status Description")

    start_row = 3
    start_col = 3  # Column C

    # Headers in C3 and D3
    legend_sheet.cell(row=start_row, column=start_col, value="Status Value")
    legend_sheet.cell(row=start_row, column=start_col + 1, value="Description")

    header_fill = PatternFill(start_color="FF284464", end_color="FF284464", fill_type="solid")  # Dark blue fill
    header_font = Font(bold=True, color="FFFFFFFF")  # White font

    for col in range(start_col, start_col + 2):
        cell = legend_sheet.cell(row=start_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = start_row + 1
    for status, desc in status_descriptions.items():
        status_cell = legend_sheet.cell(row=current_row, column=start_col, value=status)
        legend_sheet.cell(row=current_row, column=start_col + 1, value=desc)

        color_hex = status_fill_colors.get(status)
        if color_hex:
            fill = PatternFill(start_color=f"FF{color_hex}", end_color=f"FF{color_hex}", fill_type="solid")
            status_cell.fill = fill
            status_cell.font = Font(bold=False)

        current_row += 1

    legend_sheet.column_dimensions['C'].width = 13
    legend_sheet.column_dimensions['D'].width = 47

    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    max_row = start_row + len(status_descriptions)
    for row in legend_sheet.iter_rows(min_row=start_row, max_row=max_row, min_col=start_col, max_col=start_col + 1):
        for cell in row:
            cell.border = thin_border


def process_file(file_path: Path, data_sheet_name: str, bu_name: str):
    try:
        workbook = load_workbook(file_path)
        if data_sheet_name not in workbook.sheetnames:
            print(f"Sheet '{data_sheet_name}' not found in {file_path.name}, skipping.")
            return

        worksheet = workbook[data_sheet_name]
        status_col_idx = find_status_column_index(worksheet)
        if not status_col_idx:
            print(f"'Status' column not found in sheet '{data_sheet_name}' of {file_path.name}, skipping.")
            return

        clear_status_column_values(worksheet, status_col_idx)

        insert_ticket_no_column(worksheet, status_col_idx)
        ticket_col_idx = status_col_idx + 1

        insert_comments_column(worksheet, ticket_col_idx)
        comments_col_idx = ticket_col_idx + 1

        insert_it_column(worksheet, comments_col_idx)
        it_col_idx = comments_col_idx + 1

        status_col_letter = get_column_letter(status_col_idx)
        ticket_col_letter = get_column_letter(ticket_col_idx)
        comments_col_letter = get_column_letter(comments_col_idx)
        it_col_letter = get_column_letter(it_col_idx)

        add_status_data_validation_safe(worksheet, status_col_letter, allowed_status_values)
        add_ticket_no_numeric_validation_safe(worksheet, ticket_col_letter)
        add_it_data_validation_safe(worksheet, it_col_letter, bu_name)
        add_status_conditional_formatting_safe(worksheet, status_col_letter)

        add_status_legend_sheet(workbook, status_descriptions)

        new_file_path = approval_folder / file_path.name
        workbook.save(new_file_path)
        print(f"Processed and saved: {new_file_path}")

    except Exception as e:
        print(f"Error processing {file_path.name}: {e}")


if __name__ == "__main__":
    for country in countries:
        for report_type in report_types:
            file_name = f"{country}_{year}_{current_month_num}_Final_Report_{report_type}_Report.xlsx"
            file_path = Path(file_name)
            data_sheet_name = f"{country}_{report_type}_Scan_Result"

            if file_path.exists():
                process_file(file_path, data_sheet_name, country)
            else:
                print(f"File {file_path} does not exist.")
