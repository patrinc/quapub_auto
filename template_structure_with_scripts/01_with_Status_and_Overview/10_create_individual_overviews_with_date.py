import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from pathlib import Path
import copy

# Add parent directory (where config.py is located) to sys.path
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, '..'))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

import config  # Import config.py from parent directory

previous_month_num = config.previous_month_num
current_month_num = config.current_month_num
year = config.year
assessment_date = config.assessment_date  # New variable added in config.py

countries = ['Brazil', 'CSA', 'India', 'Mexico', 'SFTL']
report_types = ['Network', 'Server']

def prepare_excel_file(file_path):
    try:
        workbook = load_workbook(file_path)
        print(f"File {file_path} is ready for overview creation.")
    except Exception as e:
        print(f"Error preparing file {file_path}: {e}")

def create_individual_overview(file_path, sheet_name):
    try:
        workbook = load_workbook(file_path)
        active_sheet = workbook.active
        active_sheet.title = sheet_name
        workbook.save(file_path)
    except Exception as e:
        print(f"Error loading or renaming workbook {file_path}: {e}")
        return
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=7)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook[sheet_name]
    except Exception as e:
        print(f"Error loading workbook {file_path} for overview creation: {e}")
        return

    severity_counts = df['Severity'].value_counts().reindex(range(1, 6), fill_value=0)

    if 'Overview' in workbook.sheetnames:
        workbook.remove(workbook['Overview'])

    overview_sheet = workbook.create_sheet('Overview')
    workbook._sheets.insert(0, workbook._sheets.pop(workbook._sheets.index(overview_sheet)))

    overview_sheet.merge_cells('D5:D6')
    overview_sheet['D5'].value = 'Assessment Type'
    overview_sheet['D5'].alignment = Alignment(horizontal='center', vertical='center')
    overview_sheet['D7'].value = sheet_name

    # Use assessment_date from config.py here
    overview_sheet['E5'].value = 'Assessment Date'
    overview_sheet['E6'].value = '(DD-MM-YY)'
    overview_sheet['E7'].value = assessment_date

    overview_sheet.merge_cells('F5:J5')
    overview_sheet['F5'].value = 'Infra Vulnerabilities'
    overview_sheet['F5'].alignment = Alignment(horizontal='center', vertical='center')

    color_fills = {
        5: PatternFill(start_color='c00000', end_color='c00000', fill_type='solid'),
        4: PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid'),
        3: PatternFill(start_color='ffc000', end_color='ffc000', fill_type='solid'),
        2: PatternFill(start_color='92d050', end_color='92d050', fill_type='solid'),
        1: PatternFill(start_color='00b0f0', end_color='00b0f0', fill_type='solid')
    }

    severity_headers = ['Severity - 5 (Critical)', 'Severity - 4 (High)', 'Severity - 3 (Medium)', 'Severity - 2 (Low)', 'Severity - 1 (Info)']
    for col_num, header in enumerate(severity_headers, start=6):
        cell = overview_sheet.cell(row=6, column=col_num, value=header)
        severity_level = 5 - (col_num - 6)
        cell.fill = color_fills[severity_level]
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col_num, severity in enumerate(range(5, 0, -1), start=6):
        cell = overview_sheet.cell(row=7, column=col_num, value=severity_counts[severity])
        cell.fill = color_fills[severity]

    bold_font = Font(bold=True)
    for row in [5, 6]:
        for col in range(4, 11):
            overview_sheet.cell(row=row, column=col).font = bold_font

    overview_sheet.column_dimensions['D'].width = 20
    for col_letter in ['E', 'F', 'G', 'H', 'I', 'J']:
        overview_sheet.column_dimensions[col_letter].width = 18

    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    for row in range(5, 8):
        for col in range(4, 11):
            overview_sheet.cell(row=row, column=col).border = thin_border

    center_wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in range(5, 8):
        for col in range(4, 11):
            cell = overview_sheet.cell(row=row, column=col)
            cell.alignment = center_wrap_alignment

    try:
        workbook.save(file_path)
    except Exception as e:
        print(f"Error saving workbook {file_path}: {e}")

# Example to generate overviews for each final report file
for country in countries:
    for report_type in report_types:
        file_path = Path(f'{country}_{year}_{current_month_num}_Final_Report_with_Status_{report_type}_Report.xlsx')
        sheet_name = f'{country}_{report_type}_Scan_Result'
        if file_path.exists():
            prepare_excel_file(file_path)
            create_individual_overview(file_path, sheet_name)
        else:
            print(f'File {file_path} does not exist.')
