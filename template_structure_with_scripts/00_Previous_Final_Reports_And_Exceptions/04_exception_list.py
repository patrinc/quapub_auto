import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

directory = "."
output_folder = r"C:\Users\example.user1\OneDrive - exampledomain\General - GSRC Global\CIC\Vulnerability Mgmt\Qualys Global Exception Lists"
required_columns = ["IP", "QID", "Title", "Severity", "Status", "Ticket No.", "Comments", "IT", "SLA Status"]
cleanup_statuses = {"Inactive", "Shutdown"}
exception_statuses = {
    "Not Applicable",
    "False Positive",
    "Mitigated",
    "Risk Accepted"
}
status_fill_colors = {
    "Patched": "FFA5D6A7",
    "Unpatched": "FFEF9A9A",
    "Risk Accepted": "FFFFF59D",
    "Mitigated": "FF64B5F6",
    "In Progress": "FF90CAF9",
    "False Positive": "FFE0E0E0",
    "Not Applicable": "FF9575CD",
    "Deferred": "FFFFF176",
    "Inactive": "FFB0BEC5",
    "Shutdown": "FF78909C"
}

def extract_month_number(filename):
    match = re.search(r'_(0[1-9]|1[0-2])_', filename)
    return match.group(1) if match else None

def extract_country(filename):
    return filename.split('_')[0] if filename else "Unknown"

months_found = set()
for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        month_num = extract_month_number(filename)
        if month_num:
            months_found.add(month_num)

if not months_found:
    print("No files with month pattern found in filenames in current directory.")
    exit()

current_month_num = max(months_found)
prev_month_int = int(current_month_num) - 1
if prev_month_int == 0:
    prev_month_int = 12
previous_month_num = f"{prev_month_int:02d}"

df_list_cleanup = []
df_list_exceptions = []

for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        month_num = extract_month_number(filename)
        if month_num == current_month_num:
            filepath = os.path.join(directory, filename)
            try:
                df = pd.read_excel(filepath)
                filtered_cols = [col for col in required_columns if col in df.columns]
                df = df[filtered_cols].copy()
                country = extract_country(filename)
                df["IT"] = country
                df["Month"] = current_month_num
                df_cleanup = df[df["Status"].isin(cleanup_statuses)]
                df_exceptions = df[df["Status"].isin(exception_statuses)]
                if not df_cleanup.empty:
                    df_list_cleanup.append(df_cleanup)
                if not df_exceptions.empty:
                    df_list_exceptions.append(df_exceptions)
            except Exception as e:
                print(f"Error processing {filename}: {e}")

if not df_list_cleanup and not df_list_exceptions:
    print("No data matching cleanup or exception statuses found for current month files.")
    exit()

cleanup_df = pd.concat(df_list_cleanup, ignore_index=True) if df_list_cleanup else pd.DataFrame(columns=required_columns + ["IT", "Month"])
exceptions_df = pd.concat(df_list_exceptions, ignore_index=True) if df_list_exceptions else pd.DataFrame(columns=required_columns + ["IT", "Month"])

prev_master_file = os.path.join(output_folder, f"Global_Exception_List_{previous_month_num}.xlsx")
curr_master_file = os.path.join(output_folder, f"Global_Exception_List_{current_month_num}.xlsx")

if os.path.exists(prev_master_file):
    try:
        prev_xl = pd.ExcelFile(prev_master_file)
        prev_cleanup_df = pd.read_excel(prev_master_file, sheet_name='Cleanup List') if 'Cleanup List' in prev_xl.sheet_names else pd.DataFrame(columns=cleanup_df.columns)
        prev_exceptions_df = pd.read_excel(prev_master_file, sheet_name='Exception List') if 'Exception List' in prev_xl.sheet_names else pd.DataFrame(columns=exceptions_df.columns)
        print(f"Previous month cleanup entries count: {len(prev_cleanup_df)}")
        print(f"Previous month exception entries count: {len(prev_exceptions_df)}")
        combined_cleanup_df = pd.concat([prev_cleanup_df, cleanup_df], ignore_index=True)
        combined_exceptions_df = pd.concat([prev_exceptions_df, exceptions_df], ignore_index=True)
        print(f"New cleanup entries count: {len(cleanup_df)}")
        print(f"New exception entries count: {len(exceptions_df)}")
    except Exception as e:
        print(f"Error reading previous month file: {e}")
        combined_cleanup_df = cleanup_df
        combined_exceptions_df = exceptions_df
else:
    combined_cleanup_df = cleanup_df
    combined_exceptions_df = exceptions_df
    print(f"No previous month exception list found. Saving current month data as new file.")

with pd.ExcelWriter(curr_master_file, engine='openpyxl') as writer:
    combined_cleanup_df.to_excel(writer, sheet_name='Cleanup List', index=False)
    combined_exceptions_df.to_excel(writer, sheet_name='Exception List', index=False)

wb = load_workbook(curr_master_file)

def apply_status_colors(ws):
    status_col = None
    for col_num, cell in enumerate(ws[1], 1):
        if cell.value == "Status":
            status_col = col_num
            break
    if status_col is None:
        print(f"Status column not found in {ws.title} sheet.")
        return
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=status_col, max_col=status_col):
        cell = row[0]
        status_val = cell.value
        if status_val in status_fill_colors:
            color = status_fill_colors[status_val]
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

for sheet_name in ['Cleanup List', 'Exception List']:
    if sheet_name in wb.sheetnames:
        apply_status_colors(wb[sheet_name])
    else:
        print(f"{sheet_name} sheet not found in the saved file.")

wb.save(curr_master_file)
print(f"Saved both lists with colors to {curr_master_file}")
