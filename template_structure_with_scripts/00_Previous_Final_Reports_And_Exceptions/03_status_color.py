import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Directory containing processed Excel files
directory = "."

# Color codes for Status values
status_fill_colors = {
    "Patched": "FFA5D6A7",
    "Unpatched": "FFEF9A9A",
    "Risk Accepted": "FFFFF59D",
    "Mitigated": "FF64B5F6",
    "In Progress": "FF90CAF9",
    "False Positive": "FFE0E0E0",
    "Not Applicable": "FF9575CD",
    "Deferred": "FFFFF176",
    "Inactive": "FFB0BEC5",
    "Shutdown": "FF78909C"
}

def apply_status_colors(filepath):
    wb = load_workbook(filepath)
    
    for sheet_name in wb.sheetnames:
        if sheet_name in ['Overview', 'Status Description']:
            continue  # Skip these sheets if you want
        
        ws = wb[sheet_name]
        
        # Find the "Status" column by header in first row
        status_col = None
        header_row = None
        
        # Assume header is the first row with non-empty cells
        for row in ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                if cell.value == "Status":
                    status_col = cell.column  # this gives the column number (int)
                    header_row = cell.row
                    break
            if status_col:
                break
        
        if not status_col:
            print(f"No 'Status' column found in sheet '{sheet_name}' of {filepath}")
            continue
        
        # Apply fill colors to cells in the "Status" column starting from row after header
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=status_col, max_col=status_col):
            cell = row[0]
            status_value = cell.value
            if status_value in status_fill_colors:
                fill_color = status_fill_colors[status_value]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    wb.save(filepath)
    print(f"Applied colors and saved: {filepath}")

# Run for all .xlsx files in the directory
for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        filepath = os.path.join(directory, filename)
        print(f"Processing file: {filename}")
        apply_status_colors(filepath)
